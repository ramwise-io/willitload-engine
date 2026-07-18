"""
willitload.tier0.parsers — Specialized parsers for non-DuckDB formats.

Handles: SQLite, XML, ZIP (archive enumeration).
Excel (openpyxl) is deferred to a fast-follow release.

These parsers are the Python-side minority — called only for file types that
DuckDB doesn't natively read. They are bounded by sample-then-confirm and
never read full file contents unnecessarily.

All parsers populate a PhysicalFile in place, matching the same contract
as duckdb_reader.profile_file().
"""

from __future__ import annotations

import sqlite3
import xml.etree.ElementTree as ET
import zipfile
from pathlib import Path

from willitload.tier0.physical import PhysicalFile
from willitload.models import Bucket

# Maximum bytes to sample from XML files for structure detection
_XML_SAMPLE_BYTES = 256 * 1024
# Maximum decompressed size allowed per archive (zip-bomb protection)
_MAX_DECOMPRESSED_BYTES = 512 * 1024 * 1024  # 512 MB
# Maximum archive nesting depth scanned
MAX_ARCHIVE_DEPTH = 1


def profile_sqlite(path: Path, pf: PhysicalFile) -> None:
    """
    Profile a SQLite file: extract all table names and per-table column schemas
    from sqlite_master (no data rows read).
    """
    try:
        conn = sqlite3.connect(f"file:{path}?mode=ro", uri=True)
        cursor = conn.cursor()

        # Get all user tables
        cursor.execute(
            "SELECT name FROM sqlite_master WHERE type='table' ORDER BY name"
        )
        tables = [row[0] for row in cursor.fetchall()]

        if not tables:
            pf.bucket = Bucket.DEGRADED
            pf.error = "SQLite file has no user tables"
            conn.close()
            return

        # Profile the first table as the primary schema
        # Multi-table SQLite → column names are prefixed with table name for clarity
        all_columns: list[str] = []
        for table in tables:
            cursor.execute(f"PRAGMA table_info('{table}')")  # noqa: S608 — table name from sqlite_master, safe
            cols = cursor.fetchall()
            for col in cols:
                # col = (cid, name, type, notnull, dflt_value, pk)
                col_name = f"{table}.{col[1]}" if len(tables) > 1 else col[1]
                all_columns.append(col_name)

        conn.close()

        pf.raw_column_names = all_columns
        pf.column_count = len(all_columns)
        pf.has_header = True  # SQLite schema is always named
        pf.bucket = Bucket.PROFILED

    except sqlite3.Error as e:
        pf.bucket = Bucket.DEGRADED
        pf.error = f"SQLite read error: {e}"
    except Exception as e:
        pf.bucket = Bucket.DEGRADED
        pf.error = f"SQLite parser error: {e}"


def profile_xml(path: Path, pf: PhysicalFile) -> None:
    """
    Profile an XML file: extract the element hierarchy as a path set.
    Reads only a sample for large files; builds a set of unique element paths.
    """
    try:
        # Read limited sample
        with open(path, "rb") as fh:
            sample = fh.read(_XML_SAMPLE_BYTES)

        # Parse the sample (may be truncated — handle gracefully)
        try:
            root = ET.fromstring(sample.decode("utf-8", errors="replace"))
        except ET.ParseError:
            # Try wrapping in a synthetic root if the sample cut mid-element
            try:
                root = ET.fromstring(
                    b"<__root__>" + sample + b"</__root__>"
                )
            except ET.ParseError:
                pf.bucket = Bucket.DEGRADED
                pf.error = "XML parse error on sample"
                return

        # Extract unique element paths (breadth-first, up to depth 5)
        paths: set[str] = set()
        _collect_paths(root, "", paths, max_depth=5)

        if not paths:
            pf.bucket = Bucket.DEGRADED
            pf.error = "XML: no element paths found in sample"
            return

        pf.raw_column_names = sorted(paths)
        pf.column_count = len(paths)
        pf.has_header = True  # XML element names are always "named"
        pf.bucket = Bucket.PROFILED

    except OSError as e:
        pf.bucket = Bucket.REFUSED
        pf.error = f"XML access error: {e}"
    except Exception as e:
        pf.bucket = Bucket.DEGRADED
        pf.error = f"XML parser error: {e}"


def _collect_paths(
    element: ET.Element,
    prefix: str,
    out: set[str],
    max_depth: int,
    depth: int = 0,
) -> None:
    """Recursively collect element path strings up to max_depth."""
    if depth > max_depth:
        return
    tag = element.tag.split("}")[-1] if "}" in element.tag else element.tag  # strip namespace
    path = f"{prefix}/{tag}" if prefix else tag
    out.add(path)
    for child in element:
        _collect_paths(child, path, out, max_depth, depth + 1)


def profile_zip(
    path: Path,
    pf: PhysicalFile,
    current_depth: int = 0,
) -> list[PhysicalFile]:
    """
    Enumerate files within a ZIP archive.

    Returns a list of PhysicalFile objects for each member (not yet profiled —
    the resolver calls profile_file on each member separately).
    The containing archive's pf is marked as CATALOGUED (it's a container, not a data file).

    Zip-bomb protection: refuses to enumerate if decompressed size exceeds ceiling.
    Encrypted archives: catalogued with ENCRYPTED_ARCHIVE finding.
    """
    if current_depth >= MAX_ARCHIVE_DEPTH:
        pf.bucket = Bucket.CATALOGUED
        pf.error = f"Archive nesting depth {current_depth} exceeds scan limit ({MAX_ARCHIVE_DEPTH})"
        return []

    members: list[PhysicalFile] = []

    try:
        with zipfile.ZipFile(path, "r") as zf:
            infos = zf.infolist()

            # Check for encryption
            has_encrypted = any(info.flag_bits & 0x1 for info in infos)
            if has_encrypted:
                pf.bucket = Bucket.REFUSED
                pf.error = "Encrypted archive — contents not inspected"
                return []

            # Zip-bomb protection: check total decompressed size
            total_decompressed = sum(info.file_size for info in infos)
            if total_decompressed > _MAX_DECOMPRESSED_BYTES:
                pf.bucket = Bucket.CATALOGUED
                pf.error = (
                    f"Archive decompressed size ({total_decompressed:,} bytes) "
                    f"exceeds ceiling ({_MAX_DECOMPRESSED_BYTES:,} bytes)"
                )
                return []

            # Mark the archive itself as a container (not a data file)
            pf.bucket = Bucket.CATALOGUED
            pf.format_detected = "zip"

            # Enumerate members — return them for the resolver to profile
            for info in infos:
                if info.filename.endswith("/"):
                    continue  # skip directory entries
                member_path = path / info.filename  # synthetic path for tracking
                member_pf = PhysicalFile(
                    path=member_path,
                    size_bytes=info.file_size,
                    container_path=str(path),
                    archive_depth=current_depth + 1,
                )
                members.append(member_pf)

    except zipfile.BadZipFile as e:
        pf.bucket = Bucket.DEGRADED
        pf.error = f"Bad ZIP file: {e}"
    except Exception as e:
        pf.bucket = Bucket.DEGRADED
        pf.error = f"ZIP parser error: {e}"

    return members


def profile_excel(path: Path, pf: PhysicalFile) -> None:
    """
    Profile an Excel file (.xlsx) using openpyxl in read-only, data-only mode.
    Extracts sheet names, column headers, and infers column types from a row sample.
    """
    try:
        import openpyxl
        from willitload.types import TypeClass

        # Open workbook in read-only and data-only (resolves formulas) mode
        wb = openpyxl.load_workbook(filename=str(path), read_only=True, data_only=True)
        sheet_names = wb.sheetnames

        if not sheet_names:
            pf.bucket = Bucket.DEGRADED
            pf.error = "Excel file contains no sheets"
            wb.close()
            return

        # Profile the first sheet
        ws = wb[sheet_names[0]]

        # Read first row for headers
        row_iter = ws.iter_rows(values_only=True)
        try:
            first_row = next(row_iter, None)
        except Exception:
            first_row = None

        if not first_row:
            pf.bucket = Bucket.DEGRADED
            pf.error = "Excel sheet is empty"
            wb.close()
            return

        # Extract column headers
        raw_names = []
        for cell in first_row:
            if cell is None:
                raw_names.append("")
            else:
                raw_names.append(str(cell).strip())

        # Trim trailing empty columns
        while raw_names and raw_names[-1] == "":
            raw_names.pop()

        if not raw_names:
            pf.bucket = Bucket.DEGRADED
            pf.error = "Excel sheet has no columns"
            wb.close()
            return

        column_count = len(raw_names)

        # Sample up to 100 data rows to infer types
        sample_rows = []
        for _ in range(100):
            try:
                row = next(row_iter, None)
                if row is None:
                    break
                sample_rows.append(row[:column_count])
            except Exception:
                break

        wb.close()

        # Type inference per column based on cell values
        col_types = {}
        for col_idx in range(column_count):
            col_name = raw_names[col_idx]
            if not col_name:
                col_name = f"column{col_idx}"
                raw_names[col_idx] = col_name

            values = [
                row[col_idx]
                for row in sample_rows
                if col_idx < len(row) and row[col_idx] is not None
            ]

            from datetime import datetime, date
            types_seen = set()
            for v in values:
                if isinstance(v, bool):
                    types_seen.add(TypeClass.BOOL)
                elif isinstance(v, int):
                    types_seen.add(TypeClass.INT)
                elif isinstance(v, float):
                    types_seen.add(TypeClass.DECIMAL)
                elif isinstance(v, (datetime, date)):
                    # openpyxl returns datetime objects for date/time columns
                    types_seen.add(TypeClass.TIMESTAMP)
                else:
                    types_seen.add(TypeClass.TEXT)

            if not types_seen:
                inferred = TypeClass.ANY
            elif len(types_seen) == 1:
                inferred = list(types_seen)[0]
            else:
                if types_seen == {TypeClass.INT, TypeClass.DECIMAL}:
                    inferred = TypeClass.DECIMAL
                else:
                    inferred = TypeClass.TEXT

            col_types[col_name] = inferred

        # Populate PhysicalFile properties
        pf.raw_column_names = raw_names
        pf.column_count = column_count
        pf.has_header = True
        pf.column_types = col_types
        pf.bucket = Bucket.PROFILED

    except Exception as e:
        pf.bucket = Bucket.DEGRADED
        pf.error = f"Excel parsing error: {e}"
